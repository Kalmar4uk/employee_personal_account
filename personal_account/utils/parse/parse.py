from datetime import datetime

from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from utils.constants import MONTHS, TIME_FORMAT


def open_wb(file: any, holiday_gsma: bool | None = None) -> Workbook:

    try:
        wb = definition_wb(file=file, holiday_gsma=holiday_gsma)
    except FileNotFoundError as e:
        raise ValueError(str(e))
    except InvalidFileException as e:
        raise ValueError(str(e))

    if not holiday_gsma:
        month, year = timezone.now().strftime("%m %Y").split(" ", 1)
        period = f"{MONTHS.get(int(month))} {year}"
        try:
            return wb[period]
        except KeyError as e:
            raise ValueError(str(e))
    return wb.active


def definition_wb(file: any, holiday_gsma: bool | None = None) -> Workbook:
    try:
        if not holiday_gsma:
            return load_workbook(filename=file)
        return load_workbook(filename=file)
    except FileNotFoundError as e:
        raise ValueError(str(e))
    except InvalidFileException as e:
        raise ValueError(str(e))


def preparation_time(
        time_start, time_end
) -> dict[str, str | datetime]:

    time_start = datetime.strptime(
        time_start.rstrip(),
        TIME_FORMAT
    ).time()

    time_end = datetime.strptime(
        time_end.rstrip(),
        TIME_FORMAT
    ).time()

    return {
        "time_start": time_start,
        "time_end": time_end,
    }


def preparation_type_shift(type_shift: str) -> dict[str, str]:
    night_shift = False

    type = type_shift.strip()

    if (type := type_shift.strip()) == "Ночное дежурство":
        night_shift = True

    return {
        "type": type,
        "night_shift": night_shift
    }
